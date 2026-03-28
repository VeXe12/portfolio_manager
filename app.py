import re
import streamlit as st
import yfinance as yf
import pandas as pd
import matplotlib
matplotlib.use('Agg') # Forces Matplotlib to run in the background (Cloud-Safe)
import matplotlib.pyplot as plt
import numpy as np
import math
from datetime import datetime
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule

class PortfolioManager:
    def __init__(self, file_path):
        self.file_path = file_path
        
        # Read the file
        if file_path.name.endswith('.csv'):
            self.df = pd.read_csv(file_path)
        else:
            self.df = pd.read_excel(file_path)
            
        # --- THE DATA CLEANING STEP ---
        # 1. Remove accidental spaces from column names
        self.df.columns = self.df.columns.str.strip()
        
        # 2. Force tickers to be UPPERCASE and remove invisible spaces
        self.df['Ticker'] = self.df['Ticker'].str.upper().str.strip()
        
        # --- COLUMN VALIDATION ---
        required_columns = ['Ticker', 'Quantity', 'Buy_Price', 'Sector']
        for col in required_columns:
            if col not in self.df.columns:
                raise ValueError(f"Missing required column: '{col}'. Please check your file formatting.")
        
        self.benchmark_ticker = "^NSEI" 
        self.advice_summary = ""
        print(f"--- Portfolio Loaded: {len(self.df)} Stocks ---")

    def update_prices(self):
        tickers = " ".join(self.df['Ticker'])
        data = yf.download(tickers, period="1d")['Close']
        current_prices = data.iloc[-1]
        self.df['Current_Price'] = self.df['Ticker'].map(current_prices)
        
        if self.df['Current_Price'].isnull().any():
            self.df = self.df.dropna(subset=['Current_Price'])

    def calculate_metrics(self):
        self.df['Invested_Val'] = self.df['Quantity'] * self.df['Buy_Price']
        self.df['Current_Val'] = self.df['Quantity'] * self.df['Current_Price']
        self.df['P_L'] = self.df['Current_Val'] - self.df['Invested_Val']
        self.df['ROI'] = (self.df['P_L'] / self.df['Invested_Val']) * 100

    def generate_advice(self):
        """Analyzes performance and stores it as a string for the email."""
        advice = "--- AI Agent Analysis ---\n"
        
        # Benchmark Logic
        nifty = yf.Ticker(self.benchmark_ticker).history(period="1y")
        nifty_roi = ((nifty['Close'].iloc[-1] - nifty['Close'].iloc[0]) / nifty['Close'].iloc[0]) * 100
        total_roi = ((self.df['Current_Val'].sum() - self.df['Invested_Val'].sum()) / self.df['Invested_Val'].sum()) * 100
        
        advice += f"Your Portfolio ROI: {total_roi:.2f}%\n"
        advice += f"NIFTY 50 ROI (1Y): {nifty_roi:.2f}%\n\n"

        for _, row in self.df.iterrows():
            if row['ROI'] > 20:
                advice += f"💰 PROFIT ALERT: {row['Ticker']} is up {row['ROI']:.1f}%\n"
            elif row['ROI'] < -15:
                advice += f"📉 STOP LOSS: {row['Ticker']} is down {row['ROI']:.1f}%\n"
        
        self.advice_summary = advice # Save for email
        print(advice) # Print to terminal

    def save_plots(self):
        """Saves the pie chart as an image for the email."""
        sector_total = self.df.groupby('Sector')['Current_Val'].sum()
        plt.figure(figsize=(8, 6))
        plt.pie(sector_total, labels=sector_total.index, autopct='%1.1f%%', startangle=140)
        plt.title("Portfolio Distribution by Sector")
        plt.savefig("sector_distribution.png") # Save file
        plt.close() # Close plot to save memory

    def export_report(self):
        """Saves the result and applies professional Excel styling."""
        filename = "Portfolio_Report.xlsx"
        
        # 1. Save the raw data first
        self.df.to_excel(filename, index=False)
        
        # 2. Re-open the workbook to apply styles
        wb = load_workbook(filename)
        ws = wb.active

        # Define visual styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 3. Apply Styles & Auto-Adjust Column Width
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            for cell in col:
                # Header Row styling
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                else:
                    # center horizontally and vertically, and wrap long text
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Add thin border around every cell
                cell.border = border

                # Determine required width
                try:
                    value_len = len(str(cell.value))
                    if value_len > max_length:
                        max_length = value_len
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        # Prepare column letters and widths for row-height approximation
        header_cells = list(ws[1])
        column_letters = [get_column_letter(c.col_idx) for c in header_cells]
        col_widths = {letter: ws.column_dimensions[letter].width for letter in column_letters}

        # 4. Apply number formats
        header_map = {cell.value: cell.col_idx for cell in header_cells}

        currency_cols = ['Buy_Price', 'Current_Price', 'Invested_Val', 'Current_Val', 'P_L']
        percent_cols = ['ROI']

        for col_name in currency_cols:
            if col_name in header_map:
                letter = get_column_letter(header_map[col_name])
                for cell in ws[letter]:
                    if cell.row == 1:
                        continue
                    cell.number_format = '₹#,##0.00'

        for col_name in percent_cols:
            if col_name in header_map:
                letter = get_column_letter(header_map[col_name])
                for cell in ws[letter]:
                    if cell.row == 1:
                        continue
                    cell.number_format = '0.00"%"'

        # 5. Conditional formatting for ROI (visual cues)
        if 'ROI' in header_map:
            roi_letter = get_column_letter(header_map['ROI'])
            roi_range = f"{roi_letter}2:{roi_letter}{ws.max_row}"
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws.conditional_formatting.add(roi_range, CellIsRule(operator='greaterThan', formula=['20'], fill=green_fill))
            ws.conditional_formatting.add(roi_range, CellIsRule(operator='lessThan', formula=['-15'], fill=red_fill))

        # 6. Freeze header, enable autofilter and save
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        # 7. Approximate and set row heights based on content and column widths
        for row_idx in range(2, ws.max_row + 1):
            max_lines = 1
            for letter in column_letters:
                cell = ws[f"{letter}{row_idx}"]
                txt = "" if cell.value is None else str(cell.value)
                if '\n' in txt:
                    lines = txt.count('\n') + 1
                else:
                    width = col_widths.get(letter, 10) or 10
                    try:
                        lines = math.ceil(len(txt) / width) if len(txt) > 0 else 1
                    except Exception:
                        lines = 1
                if lines > max_lines:
                    max_lines = lines
            ws.row_dimensions[row_idx].height = max(15, max_lines * 15)

        wb.save(filename)
        print(f"✅ Formatted report ready: {filename}")

    def send_email_report(self, receiver_email, sender_email, app_password):
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = f"📈 Portfolio Report - {datetime.now().strftime('%d %b %Y')}"

        # Attach Advice text to Body
        message.attach(MIMEText(self.advice_summary, "plain"))

        # Function to attach files
        def attach_file(filename):
            with open(filename, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename= {filename}")
            message.attach(part)

        # Attach Excel and PNG
        excel_file = "Portfolio_Report.xlsx"
        attach_file(excel_file)
        attach_file("sector_distribution.png")

        # Send
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, app_password)
            server.sendmail(sender_email, receiver_email, message.as_string())
        
        print(f"📧 Success: Report sent to {receiver_email}")

    def run(self, receiver, sender, password):
        self.update_prices()
        self.calculate_metrics()
        self.generate_advice()
        self.save_plots()
        self.export_report() 
        self.send_email_report(receiver, sender, password) 

# --- STREAMLIT USER INTERFACE ---
st.set_page_config(page_title="AI Portfolio Agent", page_icon="📈")

st.title("🚀 AI Portfolio Manager & Analyzer")
st.markdown("Upload your portfolio, and our AI agent will analyze your risk, compare it to the NIFTY 50, and email you a formatted report.")
# --- USER INSTRUCTIONS ---
with st.expander("ℹ️ How to format your Excel/CSV file"):
    st.markdown("""
    Your file **must** contain these exact four column headers:
    1. **Ticker**: The stock symbol ending in `.NS` (for NSE) or `.BO` (for BSE). Example: `RELIANCE.NS`
    2. **Quantity**: Number of shares you own.
    3. **Buy_Price**: The average price you paid per share.
    4. **Sector**: The industry (e.g., IT, Energy, Banking).
    """)

# 1. UI: Inputs
uploaded_file = st.file_uploader("Drop your Excel or CSV Portfolio here", type=["xlsx", "csv"])
user_email = st.text_input("Where should we send the report?")

# --- VALIDATION HELPER FUNCTION ---
def is_valid_email(email):
    """Uses Regex to ensure the email looks like name@domain.com"""
    regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(regex, email) is not None

# 2. UI: The Execution Button
if st.button("Analyze & Send Report"):
    
    # --- DEFENSIVE PROGRAMMING (ERROR HANDLING) ---
    if uploaded_file is None:
        st.warning("⚠️ Please upload your portfolio Excel file first.")
    
    elif not (uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.csv')):
        st.error("❌ Invalid file format. Please upload a .xlsx or .csv file.")
        
    elif not user_email:
        st.warning("⚠️ Please enter an email address.")
        
    elif not is_valid_email(user_email):
        st.error(f"❌ '{user_email}' is not a valid email format. Please check for typos.")
        
    # --- IF ALL CHECKS PASS, RUN THE ENGINE ---
    else:
        with st.spinner("Agent is fetching live market data..."):
            try:
                # Initialize your class with the uploaded file
                manager = PortfolioManager(uploaded_file)
                
                # Run the pipeline
                manager.update_prices()
                manager.calculate_metrics()
                manager.generate_advice()
                manager.save_plots()
                manager.export_report() 
                
                # Fetch credentials from Streamlit Secrets
                SENDER = st.secrets["GMAIL_USER"]
                PASSWORD = st.secrets["GMAIL_PASSWORD"]
                
                manager.send_email_report(user_email, SENDER, PASSWORD)
                
                st.success(f"✅ Success! Check your inbox at {user_email}")
                st.balloons() 
                
            except Exception as e:
                # Catch any unexpected errors (like Pandas failing to read a corrupted file)
                st.error(f"🚨 An unexpected error occurred while processing: {e}")